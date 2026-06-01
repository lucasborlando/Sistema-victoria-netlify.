from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import json, shutil, subprocess, os

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EXCEL_PATH = DATA_DIR / "inventario.xlsx"
HISTORIAL_PATH = DATA_DIR / "historial_movimientos.xlsx"
VALUES_PATH = DATA_DIR / "valores_por_codigo.json"
USERS_PATH = DATA_DIR / "usuarios.json"
SHEET_NAME = "CARGO BAVE 2015"
START_ROW = 12
COL_CODIGO = 2
COL_ITEM = 3
COL_CAT1 = 6
COL_CAT2 = 7
COL_PLANTA = 8
COL_CAT3 = 9

PUBLIC_DIR = Path(os.environ.get("SV_PUBLIC_DIR", BASE_DIR.parent / "Sistema_Victoria_NETLIFY_Informativo"))
PUBLIC_DATA = PUBLIC_DIR / "data"
PUBLIC_DOWNLOADS = PUBLIC_DIR / "downloads"

def numero(valor):
    try:
        if valor is None: return 0
        return int(float(valor))
    except Exception:
        return 0

def verify_pw(password, stored):
    import hashlib
    try:
        salt, old_hash = stored.split("$", 1)
        h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000).hex()
        return h == old_hash
    except Exception:
        return False

def normalizar_rut(rut):
    return str(rut or "").replace(".", "").replace(" ", "").lower()

def leer_usuarios_web():
    if not USERS_PATH.exists():
        return []
    try:
        users = json.loads(USERS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
    permitidos = []
    for u in users:
        if not u.get("activo", True):
            continue
        perfil = str(u.get("perfil", "")).lower()
        if perfil not in {"informativo", "supervigilante"}:
            continue
        permitidos.append({
            "id": normalizar_rut(u.get("id") or u.get("rut")),
            "rut": u.get("rut", ""),
            "nombre": u.get("nombre", ""),
            "cargo": u.get("cargo", ""),
            "perfil": perfil,
            "clave_hash": u.get("clave_hash", ""),
            "activo": bool(u.get("activo", True)),
            "debe_cambiar_clave": verify_pw("12345", u.get("clave_hash", ""))
        })
    return permitidos

def cargar_valores():
    if not VALUES_PATH.exists(): return {}
    return json.loads(VALUES_PATH.read_text(encoding="utf-8"))

def leer_inventario_publico():
    if not EXCEL_PATH.exists(): return []
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    valores = cargar_valores()
    datos = []
    for row in range(START_ROW, ws.max_row + 1):
        codigo = ws.cell(row=row, column=COL_CODIGO).value
        item = ws.cell(row=row, column=COL_ITEM).value
        if codigo is None or item is None: continue
        codigo = str(codigo).strip()
        cat1 = numero(ws.cell(row=row, column=COL_CAT1).value)
        cat2 = numero(ws.cell(row=row, column=COL_CAT2).value)
        planta = numero(ws.cell(row=row, column=COL_PLANTA).value)
        cat3 = numero(ws.cell(row=row, column=COL_CAT3).value)
        total = cat1 + cat2 + planta + cat3
        valor_unitario = int(valores.get(codigo, 0) or 0)
        datos.append({"fila": row, "codigo": codigo, "item": str(item).strip(), "cat1": cat1, "cat2": cat2, "planta": planta, "cat3": cat3, "total": total, "valor_unitario": valor_unitario, "valor_total": total * valor_unitario})
    return datos

def leer_historial_publico():
    if not HISTORIAL_PATH.exists(): return []
    wb = load_workbook(HISTORIAL_PATH, data_only=True)
    ws = wb.active
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        datos.append({
            "fecha": row[0] if len(row) > 0 else "",
            "usuario": row[1] if len(row) > 1 else "",
            "rut": row[2] if len(row) > 2 else "",
            "perfil": row[3] if len(row) > 3 else "",
            "movimiento": row[4] if len(row) > 4 else "",
            "codigo": row[5] if len(row) > 5 else "",
            "item": row[6] if len(row) > 6 else "",
            "cantidad": row[7] if len(row) > 7 else "",
            "desde": row[8] if len(row) > 8 else "",
            "hacia": row[9] if len(row) > 9 else "",
            "persona": row[10] if len(row) > 10 else "",
            "oal": row[11] if len(row) > 11 else "",
            "observacion": row[12] if len(row) > 12 else "",
        })
    datos.reverse()
    return datos

def exportar_datos_publicos(push_git=False):
    PUBLIC_DATA.mkdir(parents=True, exist_ok=True)
    PUBLIC_DOWNLOADS.mkdir(parents=True, exist_ok=True)
    inv = leer_inventario_publico()
    hist = leer_historial_publico()
    (PUBLIC_DATA / "inventario.json").write_text(json.dumps(inv, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (PUBLIC_DATA / "historial.json").write_text(json.dumps(hist, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (PUBLIC_DATA / "metadata.json").write_text(json.dumps({"actualizado": datetime.now().strftime("%d-%m-%Y %H:%M:%S"), "items": len(inv), "movimientos": len(hist)}, ensure_ascii=False, indent=2), encoding="utf-8")
    (PUBLIC_DATA / "usuarios_web.json").write_text(json.dumps(leer_usuarios_web(), ensure_ascii=False, indent=2), encoding="utf-8")
    if EXCEL_PATH.exists(): shutil.copy2(EXCEL_PATH, PUBLIC_DOWNLOADS / "inventario.xlsx")
    if HISTORIAL_PATH.exists(): shutil.copy2(HISTORIAL_PATH, PUBLIC_DOWNLOADS / "historial_movimientos.xlsx")
    if push_git or os.environ.get("SV_AUTO_GIT") == "1":
        subir_a_github()

def subir_a_github():
    if not (PUBLIC_DIR / ".git").exists():
        return
    try:
        subprocess.run(["git", "add", "data", "downloads"], cwd=PUBLIC_DIR, check=True)
        msg = "Actualización automática Sistema Victoria"
        commit = subprocess.run(["git", "commit", "-m", msg], cwd=PUBLIC_DIR, text=True, capture_output=True)
        if commit.returncode != 0 and "nothing to commit" not in (commit.stdout + commit.stderr).lower():
            return
        subprocess.run(["git", "push"], cwd=PUBLIC_DIR, check=False)
    except Exception:
        pass

if __name__ == "__main__":
    exportar_datos_publicos(push_git=True)
    print("Datos públicos exportados.")
