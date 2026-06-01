
from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import datetime
import shutil
import json
import hashlib
import secrets

try:
    from exportar_netlify import exportar_datos_publicos
except Exception:
    def exportar_datos_publicos(*args, **kwargs):
        pass

app = Flask(__name__)
app.secret_key = "sistema_victoria_local_2026"

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EXCEL_PATH = DATA_DIR / "inventario.xlsx"
BACKUP_DIR = DATA_DIR / "respaldos"
HISTORIAL_PATH = DATA_DIR / "historial_movimientos.xlsx"
USERS_PATH = DATA_DIR / "usuarios.json"
VALUES_PATH = DATA_DIR / "valores_por_codigo.json"

SHEET_NAME = "CARGO BAVE 2015"
START_ROW = 12

COL_CODIGO = 2
COL_ITEM = 3
COL_CAT1 = 6
COL_CAT2 = 7
COL_PLANTA = 8
COL_CAT3 = 9
COL_TOTAL = 10

HISTORIAL_HEADERS = [
    "Fecha", "Usuario Sistema", "RUT/Puesto", "Perfil", "Movimiento",
    "Código", "Item", "Cantidad", "Desde", "Hacia", "Responsable",
    "N° OAL", "Observación"
]

PERFILES_MODIFICAN = {"usuario", "supervigilante"}
PERFILES_VALIDOS = {"usuario", "supervigilante", "informativo"}

BACKUP_DIR.mkdir(exist_ok=True)

def hash_pw(password, salt=None):
    salt = salt or secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000).hex()
    return f"{salt}${h}"

def verify_pw(password, stored):
    try:
        salt, old_hash = stored.split("$", 1)
        return hash_pw(password, salt).split("$", 1)[1] == old_hash
    except Exception:
        return False

def asegurar_usuarios():
    if USERS_PATH.exists():
        return
    users = [
        {
            "id": "20046029-4",
            "rut": "20.046.029-4",
            "nombre": "LUCAS BORLANDO REYES",
            "cargo": "CDTE PELOTÓN",
            "perfil": "supervigilante",
            "clave_hash": hash_pw("12345"),
            "activo": True
        }
    ]
    USERS_PATH.write_text(json.dumps(users, ensure_ascii=False, indent=2), encoding="utf-8")

def cargar_usuarios():
    asegurar_usuarios()
    return json.loads(USERS_PATH.read_text(encoding="utf-8"))

def guardar_usuarios(users):
    USERS_PATH.write_text(json.dumps(users, ensure_ascii=False, indent=2), encoding="utf-8")

def cargar_valores():
    if not VALUES_PATH.exists():
        return {}
    return json.loads(VALUES_PATH.read_text(encoding="utf-8"))

def guardar_valores(valores):
    VALUES_PATH.write_text(json.dumps(valores, ensure_ascii=False, indent=2), encoding="utf-8")

def usuario_actual():
    return session.get("usuario")

def requiere_login():
    return usuario_actual() is not None

def requiere_supervigilante():
    u = usuario_actual()
    return u and u.get("perfil") == "supervigilante"

def puede_modificar_inventario():
    u = usuario_actual()
    return u and u.get("perfil") in PERFILES_MODIFICAN

def asegurar_historial_formato():
    """Mantiene el historial en un orden fijo de columnas.
    También migra historiales antiguos de 9 columnas para que no aparezcan desordenados.
    """
    if not HISTORIAL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        ws.append(HISTORIAL_HEADERS)
        wb.save(HISTORIAL_PATH)
        return

    wb = load_workbook(HISTORIAL_PATH)
    ws = wb.active
    encabezados = [str(c.value or "").strip() for c in ws[1]]

    if encabezados[:len(HISTORIAL_HEADERS)] == HISTORIAL_HEADERS:
        return

    indice = {h: i for i, h in enumerate(encabezados)}
    movimientos = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        def val(nombre, default=""):
            i = indice.get(nombre)
            if i is None or i >= len(row):
                return default
            return row[i] if row[i] is not None else default

        # Soporta historial antiguo:
        # Fecha, Movimiento, Código, Item, Cantidad, Desde, Hacia, Responsable, Observación
        movimientos.append([
            val("Fecha"),
            val("Usuario Sistema"),
            val("RUT/Puesto"),
            val("Perfil"),
            val("Movimiento"),
            val("Código"),
            val("Item"),
            val("Cantidad"),
            val("Desde"),
            val("Hacia"),
            val("Responsable"),
            val("N° OAL"),
            val("Observación"),
        ])

    nuevo = Workbook()
    nws = nuevo.active
    nws.title = "Historial"
    nws.append(HISTORIAL_HEADERS)
    for mov in movimientos:
        nws.append(mov)
    nuevo.save(HISTORIAL_PATH)

def crear_respaldo():
    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    respaldo = BACKUP_DIR / f"respaldo_inventario_{fecha}.xlsx"
    if EXCEL_PATH.exists():
        shutil.copy(EXCEL_PATH, respaldo)

def numero(valor):
    try:
        if valor is None:
            return 0
        return int(float(valor))
    except Exception:
        return 0


def normalizar_rut(valor):
    """Permite ingresar RUT con puntos o sin puntos, manteniendo guion."""
    return str(valor or "").strip().upper().replace(".", "").replace(" ", "")

def normalizar_texto(valor):
    texto = str(valor or "").strip().upper()
    texto = " ".join(texto.split())
    return texto


def leer_inventario():
    if not EXCEL_PATH.exists():
        return []

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    valores = cargar_valores()
    inventario = []

    for row in range(START_ROW, ws.max_row + 1):
        codigo = ws.cell(row=row, column=COL_CODIGO).value
        item = ws.cell(row=row, column=COL_ITEM).value
        if codigo is None or item is None:
            continue

        codigo = str(codigo).strip()
        cat1 = numero(ws.cell(row=row, column=COL_CAT1).value)
        cat2 = numero(ws.cell(row=row, column=COL_CAT2).value)
        planta = numero(ws.cell(row=row, column=COL_PLANTA).value)
        cat3 = numero(ws.cell(row=row, column=COL_CAT3).value)
        total = cat1 + cat2 + planta + cat3
        valor_unitario = int(valores.get(codigo, 0) or 0)

        inventario.append({
            "fila": row,
            "codigo": codigo,
            "item": str(item).strip(),
            "cat1": cat1,
            "cat2": cat2,
            "planta": planta,
            "cat3": cat3,
            "total": total,
            "valor_unitario": valor_unitario,
            "valor_total": total * valor_unitario
        })

    return inventario

def obtener_item_por_fila(ws, fila):
    codigo = ws.cell(row=fila, column=COL_CODIGO).value
    item = ws.cell(row=fila, column=COL_ITEM).value
    if codigo is None or item is None:
        return None

    return {
        "codigo": str(codigo).strip(),
        "item": str(item).strip(),
        "cat1": numero(ws.cell(row=fila, column=COL_CAT1).value),
        "cat2": numero(ws.cell(row=fila, column=COL_CAT2).value),
        "planta": numero(ws.cell(row=fila, column=COL_PLANTA).value),
        "cat3": numero(ws.cell(row=fila, column=COL_CAT3).value),
    }

def guardar_total(ws, fila):
    total = (
        numero(ws.cell(row=fila, column=COL_CAT1).value)
        + numero(ws.cell(row=fila, column=COL_CAT2).value)
        + numero(ws.cell(row=fila, column=COL_PLANTA).value)
        + numero(ws.cell(row=fila, column=COL_CAT3).value)
    )
    ws.cell(row=fila, column=COL_TOTAL).value = total

def registrar_historial(movimiento, codigo, item, cantidad, persona, observacion, desde, hacia, oal=""):
    u = usuario_actual() or {}

    asegurar_historial_formato()
    wb = load_workbook(HISTORIAL_PATH)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
        u.get("nombre", ""),
        u.get("rut") or u.get("id", ""),
        u.get("perfil", ""),
        movimiento,
        codigo,
        item,
        cantidad,
        desde,
        hacia,
        persona,
        oal,
        observacion
    ])

    wb.save(HISTORIAL_PATH)

@app.route("/")
def index():
    if not requiere_login():
        return redirect(url_for("login"))
    return render_template("index.html", usuario=usuario_actual())

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    identificador_raw = request.form.get("identificador", "")
    identificador = normalizar_rut(identificador_raw)
    identificador_texto = normalizar_texto(identificador_raw)
    clave = request.form.get("clave", "")

    for u in cargar_usuarios():
        posibles = [
            normalizar_rut(u.get("id", "")),
            normalizar_rut(u.get("rut", "")),
            normalizar_rut(u.get("nombre", "")),
            normalizar_texto(u.get("nombre", "")),
            normalizar_rut(u.get("cargo", "")),
            normalizar_texto(u.get("cargo", ""))
        ]
        if (identificador in posibles or identificador_texto in posibles) and u.get("activo", True) and verify_pw(clave, u.get("clave_hash", "")):
            session["usuario"] = {
                "id": u.get("id"),
                "rut": u.get("rut", ""),
                "nombre": u.get("nombre"),
                "perfil": u.get("perfil"),
                "cargo": u.get("cargo", "")
            }
            return redirect(url_for("index"))

    return render_template("login.html", error="Credenciales incorrectas o usuario inactivo.")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/api/inventario")
def api_inventario():
    if not requiere_login():
        return jsonify({"error": "No autorizado"}), 401
    return jsonify(leer_inventario())

@app.route("/api/movimiento", methods=["POST"])
def api_movimiento():
    if not requiere_login():
        return jsonify({"ok": False, "mensaje": "No autorizado."}), 401
    if not puede_modificar_inventario():
        return jsonify({"ok": False, "mensaje": "Perfil informativo: solo puede visualizar y descargar reportes."}), 403

    datos = request.get_json()
    tipo = datos.get("tipo")
    fila = int(datos.get("fila"))
    cantidad = int(datos.get("cantidad"))
    persona = datos.get("persona", "")
    observacion = datos.get("observacion", "")
    origen = datos.get("origen", "")

    if cantidad <= 0:
        return jsonify({"ok": False, "mensaje": "La cantidad debe ser mayor a cero."}), 400

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    item = obtener_item_por_fila(ws, fila)

    if not item:
        return jsonify({"ok": False, "mensaje": "Item no encontrado."}), 404

    crear_respaldo()

    if tipo == "entregar":
        columnas_entrega = {"cat1": COL_CAT1, "cat2": COL_CAT2}
        nombres_entrega = {"cat1": "CAT I Nuevo", "cat2": "CAT II Usado"}

        if origen not in columnas_entrega:
            return jsonify({"ok": False, "mensaje": "Debes seleccionar si la entrega sale desde CAT I o CAT II."}), 400

        if cantidad > item[origen]:
            return jsonify({"ok": False, "mensaje": f"No hay stock suficiente en {nombres_entrega[origen]}."}), 400

        ws.cell(row=fila, column=columnas_entrega[origen]).value = item[origen] - cantidad
        ws.cell(row=fila, column=COL_PLANTA).value = item["planta"] + cantidad
        movimiento = "Entrega a personal de planta"
        desde = f"Almacén General {nombres_entrega[origen]}"
        hacia = "Personal de Planta"

    elif tipo == "recibir":
        if cantidad > item["planta"]:
            return jsonify({"ok": False, "mensaje": "No hay stock suficiente en Personal de Planta."}), 400

        ws.cell(row=fila, column=COL_PLANTA).value = item["planta"] - cantidad
        ws.cell(row=fila, column=COL_CAT2).value = item["cat2"] + cantidad
        movimiento = "Recepción desde personal"
        desde = "Personal de Planta"
        hacia = "Almacén General CAT II Usado"

    elif tipo == "excluir":
        columnas = {"cat1": COL_CAT1, "cat2": COL_CAT2, "planta": COL_PLANTA}
        nombres = {"cat1": "CAT I Nuevo", "cat2": "CAT II Usado", "planta": "Personal de Planta"}

        if origen not in columnas:
            return jsonify({"ok": False, "mensaje": "Origen no válido."}), 400

        if cantidad > item[origen]:
            return jsonify({"ok": False, "mensaje": "No hay stock suficiente en el origen seleccionado."}), 400

        ws.cell(row=fila, column=columnas[origen]).value = item[origen] - cantidad
        ws.cell(row=fila, column=COL_CAT3).value = item["cat3"] + cantidad
        movimiento = "Envío a excluidos"
        desde = nombres[origen]
        hacia = "CAT III Excluidos"

    else:
        return jsonify({"ok": False, "mensaje": "Tipo de movimiento no válido."}), 400

    guardar_total(ws, fila)
    wb.save(EXCEL_PATH)

    registrar_historial(movimiento, item["codigo"], item["item"], cantidad, persona, observacion, desde, hacia)
    exportar_datos_publicos()
    return jsonify({"ok": True, "mensaje": "Movimiento guardado correctamente."})

@app.route("/api/item/agregar", methods=["POST"])
def api_agregar_item():
    if not requiere_login():
        return jsonify({"ok": False, "mensaje": "No autorizado."}), 401
    if not puede_modificar_inventario():
        return jsonify({"ok": False, "mensaje": "Perfil informativo: solo puede visualizar y descargar reportes."}), 403

    datos = request.get_json()
    codigo = str(datos.get("codigo", "")).strip()
    item = normalizar_texto(datos.get("item", ""))
    categoria = datos.get("categoria")
    cantidad = int(datos.get("cantidad", 0))
    valor_unitario = int(datos.get("valor_unitario", 0))
    oal = str(datos.get("oal", "")).strip()
    observacion = str(datos.get("observacion", "")).strip()

    if not codigo or not item or cantidad <= 0 or not oal:
        return jsonify({"ok": False, "mensaje": "Código, item, cantidad y N° OAL son obligatorios."}), 400

    columnas_categoria = {
        "cat1": COL_CAT1,
        "cat2": COL_CAT2,
        "planta": COL_PLANTA,
        "cat3": COL_CAT3
    }

    nombres_categoria = {
        "cat1": "CAT I Nuevo",
        "cat2": "CAT II Usado",
        "planta": "Personal de Planta",
        "cat3": "CAT III Excluidos"
    }

    if categoria not in columnas_categoria:
        return jsonify({"ok": False, "mensaje": "Categoría no válida."}), 400

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    fila_existente = None
    coincidencia = ""

    for row in range(START_ROW, ws.max_row + 1):
        codigo_actual = str(ws.cell(row=row, column=COL_CODIGO).value or "").strip()
        item_actual = normalizar_texto(ws.cell(row=row, column=COL_ITEM).value or "")

        if codigo_actual == codigo:
            fila_existente = row
            coincidencia = "código"
            break

        if item_actual and item_actual == item:
            fila_existente = row
            coincidencia = "nombre"
            break

    crear_respaldo()

    if fila_existente:
        col = columnas_categoria[categoria]
        actual = numero(ws.cell(row=fila_existente, column=col).value)
        ws.cell(row=fila_existente, column=col).value = actual + cantidad
        guardar_total(ws, fila_existente)

        codigo_real = str(ws.cell(row=fila_existente, column=COL_CODIGO).value or codigo).strip()
        item_real = str(ws.cell(row=fila_existente, column=COL_ITEM).value or item).strip()

        wb.save(EXCEL_PATH)

        valores = cargar_valores()
        if valor_unitario > 0:
            valores[codigo_real] = valor_unitario
            guardar_valores(valores)

        registrar_historial(
            "Agregar stock a item existente",
            codigo_real,
            item_real,
            cantidad,
            usuario_actual().get("nombre",""),
            f"{observacion} | Coincidencia por {coincidencia}.",
            "Ingreso por OAL",
            nombres_categoria[categoria],
            oal
        )
        exportar_datos_publicos()

        return jsonify({
            "ok": True,
            "mensaje": f"El item ya existía por {coincidencia}. Se sumó la cantidad al registro existente."
        })

    row = ws.max_row + 1
    ws.cell(row=row, column=COL_CODIGO).value = codigo
    ws.cell(row=row, column=COL_ITEM).value = item
    ws.cell(row=row, column=COL_CAT1).value = cantidad if categoria == "cat1" else 0
    ws.cell(row=row, column=COL_CAT2).value = cantidad if categoria == "cat2" else 0
    ws.cell(row=row, column=COL_PLANTA).value = cantidad if categoria == "planta" else 0
    ws.cell(row=row, column=COL_CAT3).value = cantidad if categoria == "cat3" else 0

    guardar_total(ws, row)
    wb.save(EXCEL_PATH)

    valores = cargar_valores()
    valores[codigo] = valor_unitario
    guardar_valores(valores)

    registrar_historial(
        "Agregar item nuevo",
        codigo,
        item,
        cantidad,
        usuario_actual().get("nombre",""),
        observacion,
        "Nuevo registro",
        nombres_categoria[categoria],
        oal
    )
    exportar_datos_publicos()

    return jsonify({"ok": True, "mensaje": "Item nuevo agregado correctamente."})

@app.route("/api/item/eliminar", methods=["POST"])
def api_eliminar_item():
    if not requiere_login():
        return jsonify({"ok": False, "mensaje": "No autorizado."}), 401
    if not puede_modificar_inventario():
        return jsonify({"ok": False, "mensaje": "Perfil informativo: solo puede visualizar y descargar reportes."}), 403

    datos = request.get_json()
    fila = int(datos.get("fila"))
    oal = str(datos.get("oal", "")).strip()
    motivo = str(datos.get("motivo", "")).strip()

    if not oal or not motivo:
        return jsonify({"ok": False, "mensaje": "N° OAL y motivo son obligatorios."}), 400

    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    item = obtener_item_por_fila(ws, fila)

    if not item:
        return jsonify({"ok": False, "mensaje": "Item no encontrado."}), 404

    cantidad_total = item["cat1"] + item["cat2"] + item["planta"] + item["cat3"]

    crear_respaldo()
    ws.delete_rows(fila, 1)
    wb.save(EXCEL_PATH)

    valores = cargar_valores()
    if item["codigo"] in valores:
        del valores[item["codigo"]]
        guardar_valores(valores)

    registrar_historial("Eliminar item", item["codigo"], item["item"], cantidad_total, usuario_actual().get("nombre",""), motivo, "Inventario", "Eliminado", oal)
    exportar_datos_publicos()
    return jsonify({"ok": True, "mensaje": "Item eliminado correctamente."})

@app.route("/api/historial")
def api_historial():
    if not requiere_login():
        return jsonify({"error": "No autorizado"}), 401

    asegurar_historial_formato()

    wb = load_workbook(HISTORIAL_PATH)
    ws = wb.active
    historial = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        historial.append({
            "fecha": row[0] if len(row) > 0 else "",
            "usuario": row[1] if len(row) > 1 else "",
            "rut": row[2] if len(row) > 2 else "",
            "perfil": row[3] if len(row) > 3 else "",
            "movimiento": row[4] if len(row) > 4 else "",
            "codigo": row[5] if len(row) > 5 else "",
            "item": row[6] if len(row) > 6 else "",
            "cantidad": row[7] if len(row) > 7 else "",
            "desde": row[8] if len(row) > 8 else "",
            "hacia": row[9] if len(row) > 9 else "",
            "persona": row[10] if len(row) > 10 else "",
            "oal": row[11] if len(row) > 11 else "",
            "observacion": row[12] if len(row) > 12 else "",
        })

    historial.reverse()
    return jsonify(historial)

@app.route("/api/usuarios")
def api_usuarios():
    if not requiere_supervigilante():
        return jsonify({"error": "No autorizado"}), 403

    users = cargar_usuarios()
    return jsonify([{k:v for k,v in u.items() if k != "clave_hash"} for u in users])

@app.route("/api/usuarios", methods=["POST"])
def api_crear_usuario():
    if not requiere_supervigilante():
        return jsonify({"ok": False, "mensaje": "No autorizado."}), 403

    datos = request.get_json()
    rut = str(datos.get("rut", "")).strip()
    nombre = str(datos.get("nombre", "")).strip().upper()
    cargo = str(datos.get("cargo", "")).strip().upper()
    perfil = datos.get("perfil", "usuario")
    clave = str(datos.get("clave", "12345")).strip() or "12345"
    user_id = normalizar_rut(rut) or normalizar_texto(nombre)

    if not nombre or perfil not in PERFILES_VALIDOS:
        return jsonify({"ok": False, "mensaje": "Nombre y perfil son obligatorios. Perfiles válidos: usuario, supervigilante o informativo."}), 400

    users = cargar_usuarios()

    if any(normalizar_rut(u.get("id", "")) == normalizar_rut(user_id) or (rut and normalizar_rut(u.get("rut", "")) == normalizar_rut(rut)) for u in users):
        return jsonify({"ok": False, "mensaje": "El usuario ya existe."}), 400

    users.append({
        "id": user_id,
        "rut": rut,
        "nombre": nombre,
        "cargo": cargo,
        "perfil": perfil,
        "clave_hash": hash_pw(clave),
        "activo": True
    })

    guardar_usuarios(users)
    return jsonify({"ok": True, "mensaje": "Usuario creado correctamente."})

@app.route("/api/usuarios/eliminar", methods=["POST"])
def api_eliminar_usuario():
    if not requiere_supervigilante():
        return jsonify({"ok": False, "mensaje": "No autorizado."}), 403

    user_id = str(request.get_json().get("id", "")).strip()
    actual = usuario_actual().get("id")

    if user_id == actual:
        return jsonify({"ok": False, "mensaje": "No puedes eliminar tu propio usuario activo."}), 400

    users = cargar_usuarios()
    users = [u for u in users if str(u.get("id")) != user_id]
    guardar_usuarios(users)

    return jsonify({"ok": True, "mensaje": "Usuario eliminado correctamente."})

@app.route("/api/cambiar-clave", methods=["POST"])
def api_cambiar_clave():
    if not requiere_login():
        return jsonify({"ok": False, "mensaje": "No autorizado."}), 401

    datos = request.get_json()
    actual = str(datos.get("actual", ""))
    nueva = str(datos.get("nueva", ""))

    if len(nueva) < 4:
        return jsonify({"ok": False, "mensaje": "La nueva clave debe tener al menos 4 caracteres."}), 400

    users = cargar_usuarios()
    uid = usuario_actual().get("id")

    for u in users:
        if u.get("id") == uid:
            if not verify_pw(actual, u.get("clave_hash", "")):
                return jsonify({"ok": False, "mensaje": "La clave actual no corresponde."}), 400
            u["clave_hash"] = hash_pw(nueva)
            guardar_usuarios(users)
            return jsonify({"ok": True, "mensaje": "Contraseña actualizada correctamente."})

    return jsonify({"ok": False, "mensaje": "Usuario no encontrado."}), 404

@app.route("/descargar/inventario")
def descargar_inventario():
    if not requiere_login():
        return redirect(url_for("login"))
    return send_file(EXCEL_PATH, as_attachment=True)

@app.route("/descargar/historial")
def descargar_historial():
    if not requiere_login():
        return redirect(url_for("login"))

    asegurar_historial_formato()

    return send_file(HISTORIAL_PATH, as_attachment=True)

if __name__ == "__main__":
    exportar_datos_publicos()
    app.run(debug=False, host="0.0.0.0", port=5000)
